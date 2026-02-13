"""
/extensions/supplier_stock_files.py
Server Monitoring System v8.4.3
Copyright (c) 2025 Aleksandr Sukhanov
License: MIT
Supplier stock files downloader
Система мониторинга серверов
Версия: 8.4.3
Автор: Александр Суханов (c)
Лицензия: MIT
Получение файлов остатков поставщиков
"""

from __future__ import annotations

import base64
import csv
import ftplib
import json
import os
import logging
import math
import re
import shutil
import ssl
import subprocess
import tarfile
import threading
import tempfile
import zipfile
from http import cookiejar
from datetime import datetime, timedelta
from pathlib import Path, PureWindowsPath
from typing import Any, Dict, Iterable, List
from fnmatch import fnmatch
from urllib.error import HTTPError
from urllib.parse import urljoin, urlparse
from urllib.request import (
    HTTPBasicAuthHandler,
    HTTPCookieProcessor,
    HTTPPasswordMgrWithDefaultRealm,
    HTTPSHandler,
    Request,
    build_opener,
)

from config.settings import DATA_DIR
from extensions.extension_manager import extension_manager

SUPPLIER_STOCK_EXTENSION_ID = "supplier_stock_files"
_logger = logging.getLogger("supplier_stock_files")
_monitor_logger = logging.getLogger("monitoring")

DEFAULT_SUPPLIER_STOCK_CONFIG: Dict[str, Any] = {
    "resources": [],
    "smbclient_path": "",
    "archive_cleanup_days": 0,
    "ftp_ork": {
        "host": "",
        "login": "",
        "password": "",
    },
    "download": {
        "temp_dir": str(DATA_DIR / "supplier_stock" / "tmp"),
        "archive_dir": str(DATA_DIR / "supplier_stock" / "archive"),
        "reports_file": str(DATA_DIR / "supplier_stock" / "reports.jsonl"),
        "schedule": {"enabled": False, "time": "06:00"},
        "unpack_archive": False,
        "sources": [],
    },
    "mail": {
        "enabled": False,
        "temp_dir": str(DATA_DIR / "supplier_stock" / "mail_tmp"),
        "archive_dir": str(DATA_DIR / "supplier_stock" / "mail_archive"),
        "unpack_archive": False,
        "sources": [],
    },
    "processing": {
        "rules": [],
        "date_format": "%Y-%m-%d %H:%M",
    },
    "reporting": {
        "period_days": 7,
    },
}

DEFAULT_IEK_JSON_SETTINGS: Dict[str, Any] = {
    "stores": {
        "sherbinka": "47585e53-0113-11e0-8255-003048d2334c",
        "chehov": "238cf439-2a81-11ec-a958-00155d04ac08",
        "novosibirsk": "d664f8a2-7edb-11e5-80d6-00155d0047b0",
        "yasnogorsky": "aeef2063-c1e7-11d9-b0d7-00001a1a02c3",
        "dmd": "d8879932-a371-11ec-b3de-3c7c3fd5bba6",
    },
    "msk_stores": ["sherbinka", "chehov", "yasnogorsky", "dmd"],
    "nsk_store": "novosibirsk",
    "orc_stores": [
        {"key": "novosibirsk", "stor": "IE01"},
        {"key": "dmd", "stor": "IE02"},
        {"key": "yasnogorsky", "stor": "IE03"},
        {"key": "sherbinka", "stor": "IE04"},
    ],
    "prefix": "IEK\\",
    "outputs": {
        "orig": "Остатки ИЭК.xlsx",
        "msk": "МСК/Остатки МСК.xls",
        "nsk": "РЦ/Остатки РЦ.xls",
        "orc": "iek.csv",
    },
}
_LEGACY_IEK_OUTPUT_TEMPLATES = {
    "orig": "{source_id}_orig.xlsx",
    "msk": "{source_id}_msk.xls",
    "nsk": "{source_id}_nsk.xls",
    "orc": "{source_id}_orc.csv",
}

_scheduler_lock = threading.Lock()
_scheduler_started = False
_last_run_marker: str | None = None
_smbclient_missing_logged = False


def _log_processing(message: str, *args: object) -> None:
    try:
        _monitor_logger.info(message, *args)
    except Exception:
        formatted = message % args if args else message
        _monitor_logger.info(formatted)


def _merge_dicts(defaults: Dict[str, Any], current: Dict[str, Any]) -> Dict[str, Any]:
    result = dict(defaults)
    for key, value in current.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = _merge_dicts(result[key], value)
        else:
            result[key] = value
    return result


def normalize_supplier_stock_config(config: Dict[str, Any] | None) -> Dict[str, Any]:
    if not isinstance(config, dict):
        return dict(DEFAULT_SUPPLIER_STOCK_CONFIG)
    merged = _merge_dicts(DEFAULT_SUPPLIER_STOCK_CONFIG, config)
    cleanup_days = merged.get("archive_cleanup_days", 0)
    try:
        cleanup_days_value = int(str(cleanup_days).strip())
    except (TypeError, ValueError):
        cleanup_days_value = 0
    if cleanup_days_value < 0:
        cleanup_days_value = 0
    merged["archive_cleanup_days"] = cleanup_days_value
    resources = merged.get("resources", [])
    if isinstance(resources, list):
        for resource in resources:
            if isinstance(resource, dict):
                resource.setdefault("enabled", True)
    sources = merged.get("download", {}).get("sources", [])
    if isinstance(sources, list):
        for source in sources:
            if isinstance(source, dict):
                source.setdefault("enabled", True)
                source.setdefault("upload_subdir", "")
                source.setdefault("processing_mode", "table")
                if not isinstance(source.get("individual_directory"), dict):
                    source["individual_directory"] = {}
                source["individual_directory"].setdefault("enabled", False)
                source["individual_directory"].setdefault("unc_path", "")
                source["individual_directory"].setdefault("login", "")
                source["individual_directory"].setdefault("password", "")
                if source.get("processing_mode") == "iek_json":
                    source["iek_json"] = _normalize_iek_json_settings(source.get("iek_json"))
    mail_sources = merged.get("mail", {}).get("sources", [])
    if isinstance(mail_sources, list):
        for source in mail_sources:
            if isinstance(source, dict):
                source.setdefault("enabled", True)
                source.setdefault("upload_subdir", "")
                if not isinstance(source.get("individual_directory"), dict):
                    source["individual_directory"] = {}
                source["individual_directory"].setdefault("enabled", False)
                source["individual_directory"].setdefault("unc_path", "")
                source["individual_directory"].setdefault("login", "")
                source["individual_directory"].setdefault("password", "")
    processing_rules = merged.get("processing", {}).get("rules", [])
    if isinstance(processing_rules, list):
        for rule in processing_rules:
            if isinstance(rule, dict):
                rule.setdefault("enabled", True)
                rule["requires_processing"] = _normalize_requires_processing(rule.get("requires_processing", True))
    reporting = merged.get("reporting", {})
    if not isinstance(reporting, dict):
        reporting = {}
    reporting["period_days"] = _normalize_report_period_days(reporting.get("period_days", 7))
    merged["reporting"] = reporting
    return merged


def _normalize_requires_processing(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return True
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in ("false", "0", "no", "нет", "off"):
            return False
        if lowered in ("true", "1", "yes", "да", "on"):
            return True
    return bool(value)


def _normalize_report_period_days(value: Any) -> int:
    try:
        period = int(str(value).strip())
    except (TypeError, ValueError):
        return 7
    return 1 if period < 1 else period


def _normalize_iek_json_settings(value: Any) -> Dict[str, Any]:
    settings: Dict[str, Any] = {}
    if isinstance(value, dict):
        settings = dict(value)
    normalized = _merge_dicts(DEFAULT_IEK_JSON_SETTINGS, settings)
    if isinstance(settings.get("stores"), dict):
        normalized["stores"] = dict(settings["stores"])
    if isinstance(settings.get("msk_stores"), list):
        normalized["msk_stores"] = list(settings["msk_stores"])
    if isinstance(settings.get("orc_stores"), list):
        normalized["orc_stores"] = list(settings["orc_stores"])
    if isinstance(settings.get("prefix"), str):
        normalized["prefix"] = settings["prefix"]
    if isinstance(settings.get("nsk_store"), str):
        normalized["nsk_store"] = settings["nsk_store"]
    if isinstance(settings.get("outputs"), dict):
        normalized["outputs"] = dict(settings["outputs"])
    if not isinstance(normalized.get("stores"), dict):
        normalized["stores"] = dict(DEFAULT_IEK_JSON_SETTINGS["stores"])
    if not isinstance(normalized.get("msk_stores"), list):
        normalized["msk_stores"] = list(DEFAULT_IEK_JSON_SETTINGS["msk_stores"])
    if not isinstance(normalized.get("orc_stores"), list):
        normalized["orc_stores"] = list(DEFAULT_IEK_JSON_SETTINGS["orc_stores"])
    if not isinstance(normalized.get("outputs"), dict):
        normalized["outputs"] = dict(DEFAULT_IEK_JSON_SETTINGS["outputs"])
    else:
        outputs = dict(normalized["outputs"])
        for key, legacy_template in _LEGACY_IEK_OUTPUT_TEMPLATES.items():
            if outputs.get(key) == legacy_template:
                outputs[key] = DEFAULT_IEK_JSON_SETTINGS["outputs"][key]
        normalized["outputs"] = outputs
    return normalized


def get_supplier_stock_config() -> Dict[str, Any]:
    config = extension_manager.load_extension_config(SUPPLIER_STOCK_EXTENSION_ID)
    return normalize_supplier_stock_config(config)


def save_supplier_stock_config(config: Dict[str, Any]) -> tuple[bool, str]:
    normalized = normalize_supplier_stock_config(config)
    temp_dir = Path(normalized["download"]["temp_dir"])
    temp_dir.mkdir(parents=True, exist_ok=True)
    archive_dir = Path(normalized["download"]["archive_dir"])
    archive_dir.mkdir(parents=True, exist_ok=True)
    mail_temp_dir = Path(normalized["mail"]["temp_dir"])
    mail_temp_dir.mkdir(parents=True, exist_ok=True)
    mail_archive_dir = Path(normalized["mail"]["archive_dir"])
    mail_archive_dir.mkdir(parents=True, exist_ok=True)
    return extension_manager.save_extension_config(SUPPLIER_STOCK_EXTENSION_ID, normalized)


def _ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _render_template(value: str, now: datetime, extra_context: Dict[str, Any] | None = None) -> str:
    context = {
        "date": now.strftime("%d.%m.%Y"),
        "date_iso": now.strftime("%Y-%m-%d"),
        "datetime": now.strftime("%Y-%m-%d_%H-%M-%S"),
        "time": now.strftime("%H-%M-%S"),
    }
    if extra_context:
        context.update(extra_context)
    def _replace_date(match: re.Match[str]) -> str:
        fmt = match.group("format") or ""
        fmt = fmt.replace("\\'", "'")
        try:
            return now.strftime(fmt)
        except Exception:
            return match.group(0)

    rendered = re.sub(r"\$\(\s*date\s+'(?P<format>[^']+)'\s*\)", _replace_date, value)
    try:
        rendered = rendered.format_map(context)
    except Exception:
        pass
    rendered = re.sub(
        r"\$\{(?P<key>[A-Za-z_][A-Za-z0-9_]*)\}",
        lambda m: str(context.get(m.group("key"), m.group(0))),
        rendered,
    )
    return re.sub(
        r"\$(?P<key>[A-Za-z_][A-Za-z0-9_]*)",
        lambda m: str(context.get(m.group("key"), m.group(0))),
        rendered,
    )


def _render_iek_output_template(value: str, now: datetime, source_id: str | None) -> str:
    context = {"source_id": source_id or ""}
    return _render_template(value, now, context)


def _detect_output_format(filename: str) -> str:
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix in ("xls", "xlsx", "csv"):
        return suffix
    return "csv"


def _build_render_context(source: Dict[str, Any], now: datetime) -> Dict[str, Any]:
    context: Dict[str, Any] = {}
    vars_map = source.get("vars", {})
    if isinstance(vars_map, dict):
        for key, value in vars_map.items():
            context[key] = _render_template(str(value), now, context)
    return context


def _is_ascii(value: str) -> bool:
    try:
        value.encode("ascii")
    except UnicodeEncodeError:
        return False
    return True


def _open_with_auth_retry(
    opener: Any,
    url: str,
    headers: Dict[str, str],
    timeout: int,
    username: str,
    password: str,
    handlers: list,
) -> Any:
    request = Request(url, headers=headers)
    try:
        return opener.open(request, timeout=timeout)
    except HTTPError as exc:
        if exc.code != 401 or not (username or password):
            raise
    auth_headers = dict(headers)
    if "Authorization" not in auth_headers:
        credentials = f"{username}:{password}".encode("utf-8")
        auth_headers["Authorization"] = "Basic " + base64.b64encode(credentials).decode("ascii")
    if _is_ascii(f"{username}{password}"):
        password_mgr = HTTPPasswordMgrWithDefaultRealm()
        password_mgr.add_password(None, url, username, password)
        auth_handler = HTTPBasicAuthHandler(password_mgr)
        auth_opener = build_opener(*handlers, auth_handler)
        request = Request(url, headers=auth_headers)
        return auth_opener.open(request, timeout=timeout)
    request = Request(url, headers=auth_headers)
    return opener.open(request, timeout=timeout)


def _download_http(
    source: Dict[str, Any],
    output_path: Path,
    now: datetime,
    render_context: Dict[str, Any],
) -> Dict[str, Any]:
    timeout = int(source.get("timeout", 300))
    headers = dict(source.get("headers") or {})
    verify_ssl = bool(source.get("verify_ssl", True))
    auth = source.get("auth") or {}
    username = auth.get("username") or ""
    password = auth.get("password") or ""

    handlers = []
    cookie_jar = cookiejar.CookieJar()
    cookie_handler = HTTPCookieProcessor(cookie_jar)
    handlers.append(cookie_handler)
    if username or password:
        auth_encoding = str(auth.get("encoding", "utf-8"))
        credentials = f"{username}:{password}".encode(auth_encoding)
        headers["Authorization"] = "Basic " + base64.b64encode(credentials).decode("ascii")
    if not verify_ssl:
        handlers.append(HTTPSHandler(context=ssl._create_unverified_context()))

    opener = build_opener(*handlers)
    pre_request = source.get("pre_request")
    if pre_request:
        pre_result = _run_pre_request(pre_request, opener, now, render_context, timeout)
        if not pre_result.get("success"):
            return pre_result

    discover = source.get("discover")
    if discover:
        discover_result = _discover_url(discover, opener, now, render_context, timeout)
        if not discover_result.get("success"):
            return discover_result
        url = discover_result.get("url", "")
    else:
        url = _render_template(str(source.get("url", "")), now, render_context)

    try:
        response = _open_with_auth_retry(
            opener,
            url,
            headers,
            timeout,
            username,
            password,
            handlers,
        )
    except Exception as exc:
        return {"success": False, "error": str(exc)}

    try:
        _ensure_parent(output_path)
        with response:
            data = response.read()
        include_headers = bool(source.get("include_headers"))
        append_mode = bool(source.get("append"))
        if include_headers:
            version_map = {10: "1.0", 11: "1.1"}
            http_version = version_map.get(getattr(response, "version", 11), "1.1")
            status_line = f"HTTP/{http_version} {response.status} {response.reason}\r\n"
            header_lines = [f"{key}: {value}\r\n" for key, value in response.getheaders()]
            header_block = status_line + "".join(header_lines) + "\r\n"
            data = header_block.encode("utf-8") + data
        write_mode = "ab" if append_mode else "wb"
        with output_path.open(write_mode) as file:
            file.write(data)
        return {
            "success": True,
            "bytes": output_path.stat().st_size,
            "path": str(output_path),
        }
    finally:
        try:
            response.close()
        except Exception:
            pass


def _run_pre_request(
    pre_request: Dict[str, Any],
    opener: Any,
    now: datetime,
    render_context: Dict[str, Any],
    timeout: int,
) -> Dict[str, Any]:
    if not isinstance(pre_request, dict):
        return {"success": False, "error": "Некорректные данные pre_request"}
    url = _render_template(str(pre_request.get("url", "")), now, render_context)
    if not url:
        return {"success": False, "error": "pre_request: URL не задан"}
    method = str(pre_request.get("method", "POST")).upper()
    headers = dict(pre_request.get("headers") or {})
    data_value = pre_request.get("data")
    data_bytes: bytes | None
    if data_value is None:
        data_bytes = b"" if method in ("POST", "PUT", "PATCH") else None
    else:
        data_text = _render_template(str(data_value), now, render_context)
        data_bytes = data_text.encode("utf-8")
    if data_bytes is not None and "Content-Type" not in headers:
        headers["Content-Type"] = "application/x-www-form-urlencoded"

    max_redirects = 3
    current_url = url
    current_method = method
    current_data = data_bytes
    for _ in range(max_redirects + 1):
        request = Request(current_url, data=current_data, headers=headers, method=current_method)
        try:
            response = opener.open(request, timeout=timeout)
            with response:
                response.read()
            return {"success": True}
        except HTTPError as exc:
            if exc.code in (301, 302, 303, 307, 308) and exc.headers.get("Location"):
                current_url = urljoin(current_url, exc.headers["Location"])
                if exc.code == 303:
                    current_method = "GET"
                    current_data = None
                continue
            return {"success": False, "error": f"pre_request: {exc}"}
        except Exception as exc:
            return {"success": False, "error": f"pre_request: {exc}"}
    return {"success": False, "error": "pre_request: слишком много перенаправлений"}


def _discover_url(
    discover: Dict[str, Any],
    opener: Any,
    now: datetime,
    render_context: Dict[str, Any],
    timeout: int,
) -> Dict[str, Any]:
    if not isinstance(discover, dict):
        return {"success": False, "error": "Некорректные данные discover"}
    source_url = _render_template(str(discover.get("url", "")), now, render_context)
    pattern = _render_template(str(discover.get("pattern", "")), now, render_context)
    prefix = _render_template(str(discover.get("prefix", "")), now, render_context)
    if not source_url or not pattern:
        return {"success": False, "error": "discover: URL или шаблон не задан"}
    try:
        request = Request(source_url)
        with opener.open(request, timeout=timeout) as response:
            page_content = response.read().decode("utf-8", errors="ignore")
    except Exception as exc:
        return {"success": False, "error": f"discover: {exc}"}

    match = re.search(pattern, page_content)
    if not match:
        return {"success": False, "error": "discover: ссылка не найдена"}
    discovered = match.group(0)
    if prefix:
        discovered = prefix + discovered
    return {"success": True, "url": discovered}


def _run_shell_command(
    source: Dict[str, Any],
    now: datetime,
    temp_dir: Path,
    render_context: Dict[str, Any],
) -> Dict[str, Any]:
    import subprocess

    command = _render_template(str(source.get("command", "")), now, render_context)
    if not command:
        return {"success": False, "error": "Команда не задана"}

    result = subprocess.run(command, shell=True, capture_output=True, text=True)
    output_name = _render_template(str(source.get("output_name", "")), now)
    output_path = str(temp_dir / output_name) if output_name else ""

    return {
        "success": result.returncode == 0,
        "exit_code": result.returncode,
        "stdout": result.stdout.strip(),
        "stderr": result.stderr.strip(),
        "path": output_path,
    }


def append_supplier_stock_report(entry: Dict[str, Any]) -> None:
    config = get_supplier_stock_config()
    reports_file = Path(config["download"]["reports_file"])
    _ensure_parent(reports_file)
    with reports_file.open("a", encoding="utf-8") as file:
        file.write(json.dumps(entry, ensure_ascii=False) + "\n")


def _load_supplier_stock_reports() -> List[Dict[str, Any]]:
    config = get_supplier_stock_config()
    reports_file = Path(config["download"]["reports_file"])
    if not reports_file.exists():
        return []

    entries: List[Dict[str, Any]] = []
    with reports_file.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return entries

def _parse_report_timestamp(entry: Dict[str, Any]) -> datetime | None:
    raw = entry.get("timestamp")
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw
    try:
        return datetime.fromisoformat(str(raw))
    except ValueError:
        return None


def _filter_supplier_stock_reports(
    entries: List[Dict[str, Any]],
    period_days: int | None = None,
    source_id: str | None = None,
    source_kind: str | None = None,
) -> List[Dict[str, Any]]:
    filtered = entries
    if source_id:
        filtered = [entry for entry in filtered if str(entry.get("source_id")) == str(source_id)]
    if source_kind:
        filtered = [entry for entry in filtered if str(entry.get("source_kind")) == str(source_kind)]
    if period_days and period_days > 0:
        cutoff = datetime.now() - timedelta(days=period_days)
        filtered = [
            entry for entry in filtered
            if (_parse_report_timestamp(entry) or datetime.min) >= cutoff
        ]
    return filtered


def get_supplier_stock_reports(
    limit: int | None = 20,
    period_days: int | None = None,
    source_id: str | None = None,
    source_kind: str | None = None,
) -> List[Dict[str, Any]]:
    entries = _load_supplier_stock_reports()
    filtered = _filter_supplier_stock_reports(entries, period_days, source_id, source_kind)
    if limit is None:
        return filtered[::-1]
    return filtered[-limit:][::-1]


def get_supplier_stock_reports_total() -> int:
    config = get_supplier_stock_config()
    reports_file = Path(config["download"]["reports_file"])
    if not reports_file.exists():
        return 0

    total = 0
    with reports_file.open("r", encoding="utf-8") as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                json.loads(line)
            except json.JSONDecodeError:
                continue
            total += 1

    return total


def _map_stage_status(value: str | None) -> Dict[str, str]:
    status = (value or "").lower()
    if status in ("success", "ok", "done"):
        return {"status": "success", "icon": "🟢"}
    if status in ("error", "failed", "fail"):
        return {"status": "error", "icon": "🔴"}
    if status in ("skipped", "skip", "warning"):
        return {"status": "warning", "icon": "🟡"}
    return {"status": "unknown", "icon": "⚪️"}


def _extract_processing_status(processing: Dict[str, Any] | None) -> str | None:
    if not isinstance(processing, dict):
        return None
    status = processing.get("status")
    if status:
        return str(status)
    if processing.get("results") or processing.get("rules") or processing.get("processor"):
        return "success"
    return None


def _extract_transfer_status(processing: Dict[str, Any] | None) -> str | None:
    if not isinstance(processing, dict):
        return None
    transfer = processing.get("transfer")
    if not isinstance(transfer, dict):
        return None
    status = transfer.get("status")
    return str(status) if status else None


def summarize_supplier_stock_reports(period_days: int | None = None) -> Dict[str, List[Dict[str, Any]]]:
    entries = get_supplier_stock_reports(limit=None, period_days=period_days)
    grouped: Dict[str, Dict[str, Dict[str, Any]]] = {
        "download": {},
        "mail": {},
    }
    for entry in entries:
        source_id = entry.get("source_id") or entry.get("source_name") or "unknown"
        source_kind = entry.get("source_kind") or "download"
        if source_kind not in grouped:
            grouped[source_kind] = {}
        if source_id in grouped[source_kind]:
            continue
        processing = entry.get("processing") if isinstance(entry.get("processing"), dict) else None
        receive_status = _map_stage_status(str(entry.get("status") or ""))
        processing_status = _map_stage_status(_extract_processing_status(processing))
        transfer_status = _map_stage_status(_extract_transfer_status(processing))
        grouped[source_kind][source_id] = {
            "source_id": source_id,
            "source_name": entry.get("source_name") or source_id,
            "source_kind": source_kind,
            "method": entry.get("method"),
            "timestamp": entry.get("timestamp"),
            "receive": receive_status,
            "processing": processing_status,
            "transfer": transfer_status,
        }
    return {kind: list(items.values()) for kind, items in grouped.items()}


def build_supplier_stock_source_stats(
    source_id: str,
    source_kind: str | None,
    period_days: int | None = None,
    limit: int = 50,
) -> Dict[str, Any]:
    entries = get_supplier_stock_reports(
        limit=None,
        period_days=period_days,
        source_id=source_id,
        source_kind=source_kind,
    )
    summary = {
        "total": len(entries),
        "receive_success": 0,
        "receive_error": 0,
        "processing_success": 0,
        "processing_error": 0,
        "transfer_success": 0,
        "transfer_error": 0,
    }
    for entry in entries:
        processing = entry.get("processing") if isinstance(entry.get("processing"), dict) else None
        receive_status = _map_stage_status(str(entry.get("status") or ""))
        processing_status = _map_stage_status(_extract_processing_status(processing))
        transfer_status = _map_stage_status(_extract_transfer_status(processing))
        if receive_status["status"] == "success":
            summary["receive_success"] += 1
        if receive_status["status"] == "error":
            summary["receive_error"] += 1
        if processing_status["status"] == "success":
            summary["processing_success"] += 1
        if processing_status["status"] == "error":
            summary["processing_error"] += 1
        if transfer_status["status"] == "success":
            summary["transfer_success"] += 1
        if transfer_status["status"] == "error":
            summary["transfer_error"] += 1
    detailed = []
    limited_entries = entries[:limit]
    for entry in limited_entries:
        processing = entry.get("processing") if isinstance(entry.get("processing"), dict) else None
        receive_status = _map_stage_status(str(entry.get("status") or ""))
        processing_status = _map_stage_status(_extract_processing_status(processing))
        transfer_status = _map_stage_status(_extract_transfer_status(processing))
        detailed.append(
            {
                "timestamp": entry.get("timestamp"),
                "receive": receive_status,
                "processing": processing_status,
                "transfer": transfer_status,
                "path": entry.get("path"),
                "error": entry.get("error"),
            }
        )
    return {
        "summary": summary,
        "entries": detailed,
    }


def process_supplier_stock_file(
    file_path: str | Path,
    source_id: str | None = None,
    source_kind: str | None = None,
    input_index: int | None = None,
    original_path: str | Path | None = None,
    now: datetime | None = None,
) -> Dict[str, Any] | None:
    config = get_supplier_stock_config()
    path = Path(file_path)
    return _process_supplier_stock_file(
        path,
        config,
        now or datetime.now(),
        source_id,
        source_kind,
        input_index,
        Path(original_path) if original_path else None,
    )


def run_supplier_stock_fetch() -> Dict[str, Any]:
    config = get_supplier_stock_config()
    download_config = config.get("download", {})
    temp_dir = Path(download_config.get("temp_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["download"]["temp_dir"]))
    temp_dir.mkdir(parents=True, exist_ok=True)
    archive_dir = Path(download_config.get("archive_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["download"]["archive_dir"]))
    archive_dir.mkdir(parents=True, exist_ok=True)

    sources = download_config.get("sources", [])
    now = datetime.now()
    results: List[Dict[str, Any]] = []
    def _log(message: str, *args: object) -> None:
        try:
            _monitor_logger.info(message, *args)
        except Exception:
            formatted = message % args if args else message
            _monitor_logger.info(formatted)

    _log("📦 Остатки поставщиков: источников к обработке %s", len(sources))

    for source in sources:
        source_id = source.get("id") or source.get("name") or "unknown"
        name = source.get("name") or source_id
        if not source.get("enabled", True):
            _log("📦 Остатки поставщиков: %s пропущен (выключен)", source_id)
            results.append({"source_id": source_id, "source_name": name, "status": "skipped"})
            continue

        entry: Dict[str, Any] = {
            "timestamp": now.isoformat(),
            "source_id": source_id,
            "source_name": name,
            "method": source.get("method", "http"),
            "source_kind": "download",
        }

        try:
            render_context = _build_render_context(source, now)
            output_name = source.get("output_name")
            rendered_url = _render_template(str(source.get("url", "")), now, render_context)
            if not rendered_url:
                entry.update({"status": "error", "error": "URL не задан"})
                append_supplier_stock_report(entry)
                results.append(entry)
                _log("📦 Остатки поставщиков: %s -> error (URL не задан)", entry["source_id"])
                continue
            if output_name:
                output_name = _render_template(str(output_name), now, render_context)
                output_path = temp_dir / output_name
            else:
                output_path = temp_dir / f"{source_id}_orig"
            original_path = output_path

            entry.update({"url": rendered_url, "output_name": output_name})
            _log("📦 Остатки поставщиков: %s -> старт (%s)", entry["source_id"], rendered_url)

            if entry["method"] == "shell":
                result = _run_shell_command(source, now, temp_dir, render_context)
            else:
                result = _download_http(source, output_path, now, render_context)

            entry.update(result)
            entry["status"] = "success" if result.get("success") else "error"

            if entry["status"] == "success" and output_path.exists():
                if source.get("unpack_archive"):
                    unpacked_path = unpack_archive_file(output_path)
                    if unpacked_path:
                        entry["path"] = str(unpacked_path)
                        entry["unpacked_path"] = str(unpacked_path)
                        output_path = unpacked_path
                        _log("📦 Остатки поставщиков: %s распакован в %s", entry["source_id"], unpacked_path)
                if source.get("processing_mode") == "iek_json":
                    processing_result = _process_iek_json_file(
                        output_path,
                        config,
                        now,
                        source_id,
                        "download",
                        original_path,
                    )
                else:
                    processing_result = _process_supplier_stock_file(
                        output_path,
                        config,
                        now,
                        source_id,
                        "download",
                        1,
                        original_path,
                    )
                if processing_result:
                    entry["processing"] = processing_result
                else:
                    archive_path = _archive_original_file(original_path, archive_dir, now)
                    if archive_path:
                        entry["archive_path"] = str(archive_path)
        except Exception as exc:
            entry.update({"status": "error", "error": str(exc)})

        append_supplier_stock_report(entry)
        results.append(entry)
        if entry.get("status") == "error" and entry.get("error"):
            _log(
                "📦 Остатки поставщиков: %s -> error (%s)",
                entry["source_id"],
                entry["error"],
            )
        else:
            _log("📦 Остатки поставщиков: %s -> %s", entry["source_id"], entry["status"])

    success_count = sum(1 for item in results if item.get("status") == "success")
    cleanup_supplier_stock_archives(config, now)
    return {
        "total": len(results),
        "success": success_count,
        "failed": len(results) - success_count,
        "results": results,
    }


def _should_run_schedule(schedule: Dict[str, Any], now: datetime) -> bool:
    if not schedule.get("enabled"):
        return False

    schedule_time = str(schedule.get("time", ""))
    if not schedule_time:
        return False

    try:
        hours, minutes = [int(part) for part in schedule_time.split(":", 1)]
    except ValueError:
        return False

    if now.hour != hours or now.minute != minutes:
        return False

    global _last_run_marker
    current_date = now.strftime("%Y-%m-%d")
    marker = f"{current_date}|{schedule_time}"
    if _last_run_marker == marker:
        return False

    _last_run_marker = marker
    return True


def start_supplier_stock_scheduler() -> None:
    global _scheduler_started
    with _scheduler_lock:
        if _scheduler_started:
            return
        _scheduler_started = True

    def _loop() -> None:
        _log_processing("⏰ Запуск планировщика загрузки остатков поставщиков")
        while True:
            now = datetime.now()
            schedule = get_supplier_stock_config().get("download", {}).get("schedule", {})
            if _should_run_schedule(schedule, now):
                _log_processing("📦 Плановая загрузка остатков поставщиков")
                run_supplier_stock_fetch()
            threading.Event().wait(30)

    threading.Thread(target=_loop, daemon=True).start()


def _archive_original_file(source_path: Path, archive_dir: Path, now: datetime) -> Path | None:
    if not source_path.exists():
        return None
    archive_dir.mkdir(parents=True, exist_ok=True)
    date_stamp = now.strftime("%Y-%m-%d_%H-%M-%S")
    if source_path.suffix:
        archive_name = f"{source_path.stem}_{date_stamp}{source_path.suffix}"
    else:
        archive_name = f"{source_path.name}_{date_stamp}"
    archive_path = _unique_archive_path(archive_dir / archive_name)
    try:
        shutil.move(str(source_path), str(archive_path))
        return archive_path
    except Exception:
        return None


def archive_supplier_stock_original(source_path: Path, archive_dir: Path, now: datetime) -> Path | None:
    return _archive_original_file(source_path, archive_dir, now)


def _unique_archive_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    index = 2
    while True:
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def _parse_archive_cleanup_days(config: Dict[str, Any]) -> int:
    raw_value = config.get("archive_cleanup_days", 0)
    try:
        days = int(str(raw_value).strip())
    except (TypeError, ValueError):
        return 0
    return max(days, 0)


def cleanup_supplier_stock_archives(
    config: Dict[str, Any] | None = None,
    now: datetime | None = None,
) -> Dict[str, int]:
    config = config or get_supplier_stock_config()
    cleanup_days = _parse_archive_cleanup_days(config)
    if cleanup_days <= 0:
        return {"removed": 0, "errors": 0, "days": cleanup_days}

    current_time = now or datetime.now()
    cutoff = current_time - timedelta(days=cleanup_days)
    removed = 0
    errors = 0

    archive_dirs = [
        config.get("download", {}).get("archive_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["download"]["archive_dir"]),
        config.get("mail", {}).get("archive_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["mail"]["archive_dir"]),
    ]

    for archive_dir in archive_dirs:
        if not archive_dir:
            continue
        archive_path = Path(archive_dir)
        if not archive_path.exists():
            continue
        for item in archive_path.rglob("*"):
            if not item.is_file():
                continue
            try:
                modified = datetime.fromtimestamp(item.stat().st_mtime)
            except Exception:
                errors += 1
                continue
            if modified <= cutoff:
                try:
                    item.unlink()
                    removed += 1
                except Exception:
                    errors += 1

    if removed or errors:
        _log_processing(
            "🗑️ Очистка архива остатков: удалено %s файлов, ошибок %s",
            removed,
            errors,
        )

    return {"removed": removed, "errors": errors, "days": cleanup_days}


def _process_supplier_stock_file(
    file_path: Path,
    config: Dict[str, Any],
    now: datetime,
    source_id: str | None = None,
    source_kind: str | None = None,
    input_index: int | None = None,
    original_path: Path | None = None,
) -> Dict[str, Any] | None:
    source_name = _resolve_source_name(config, source_id, source_kind)
    processing = config.get("processing", {})
    rules = processing.get("rules", [])
    if not rules:
        _log_processing(
            "🧩 Остатки поставщиков: правила обработки не настроены для %s, выполняем выгрузку оригинала",
            file_path.name,
        )
        transfer_result = _transfer_processed_outputs(
            [],
            config,
            source_id,
            source_kind,
            original_path,
            now,
        )
        return {
            "rules": [],
            "results": [],
            "transfer": transfer_result,
            "source_id": source_id,
            "source_kind": source_kind,
            "file": str(file_path),
            "status": "skipped",
            "reason": "no_rules",
        }

    _log_processing(
        "🧩 Остатки поставщиков: проверка правил обработки для %s (источник=%s, тип=%s)",
        file_path.name,
        source_id or "не указан",
        source_kind or "не указан",
    )

    matched_rules = []
    results = []
    candidate_rules = [rule for rule in rules if isinstance(rule, dict)]
    if source_id:
        candidate_rules = [
            rule for rule in candidate_rules
            if not rule.get("source_id") or str(rule.get("source_id")) == str(source_id)
        ]
    if source_kind:
        candidate_rules = [
            rule for rule in candidate_rules
            if _processing_rule_matches_kind(rule, source_kind, config)
        ]
    active_rules = [rule for rule in candidate_rules if rule.get("active")]
    if active_rules:
        candidate_rules = active_rules
    for rule in candidate_rules:
        if not isinstance(rule, dict):
            continue
        if not rule.get("enabled", True):
            continue
        if not _processing_rule_matches(rule, file_path, input_index):
            continue
        rule_label = rule.get("id") or rule.get("name") or "rule"
        matched_rules.append(rule_label)
        _log_processing(
            "🧩 Остатки поставщиков: запуск правила обработки %s для %s",
            rule_label,
            file_path.name,
        )
        try:
            result = _run_processing_rule(file_path, rule, processing, now, input_index, source_name)
            _log_processing(
                "🧩 Остатки поставщиков: результат правила %s для %s -> %s",
                rule_label,
                file_path.name,
                result.get("status"),
            )
            results.append(result)
        except Exception as exc:
            _logger.error("⚠️ Ошибка обработки %s: %s", file_path, exc)
            _log_processing(
                "🧩 Остатки поставщиков: ошибка правила %s для %s -> %s",
                rule_label,
                file_path.name,
                exc,
            )
            results.append({"status": "error", "error": str(exc), "rule_id": rule.get("id")})

    if not matched_rules:
        _log_processing(
            "🧩 Остатки поставщиков: подходящие правила обработки не найдены для %s",
            file_path.name,
        )
        transfer_result = _transfer_processed_outputs(
            [],
            config,
            source_id,
            source_kind,
            original_path,
            now,
        )
        return {
            "rules": [],
            "results": [],
            "transfer": transfer_result,
            "source_id": source_id,
            "source_kind": source_kind,
            "file": str(file_path),
            "status": "skipped",
            "reason": "no_matching_rules",
        }

    transfer_result = _transfer_processed_outputs(
        results,
        config,
        source_id,
        source_kind,
        original_path,
        now,
    )

    return {
        "rules": matched_rules,
        "results": results,
        "transfer": transfer_result,
        "source_id": source_id,
        "source_kind": source_kind,
        "file": str(file_path),
    }


def _process_iek_json_file(
    file_path: Path,
    config: Dict[str, Any],
    now: datetime,
    source_id: str | None,
    source_kind: str | None,
    original_path: Path | None,
) -> Dict[str, Any] | None:
    source = _resolve_source_entry(config, source_id, source_kind)
    iek_settings = _normalize_iek_json_settings((source or {}).get("iek_json"))
    store_map = iek_settings.get("stores", {})
    if not store_map:
        return {"status": "error", "error": "iek_json: stores не заданы"}

    try:
        payload = json.loads(file_path.read_text(encoding="utf-8", errors="ignore"))
    except json.JSONDecodeError as exc:
        return {"status": "error", "error": f"iek_json: invalid_json ({exc})"}

    items = payload.get("shopItems", [])
    if not isinstance(items, list):
        return {"status": "error", "error": "iek_json: shopItems не найден"}

    def _get_residue(row: dict, key: str) -> int:
        store_id = store_map.get(key)
        if not store_id:
            return 0
        residues = row.get("residues", {}) if isinstance(row, dict) else {}
        value = residues.get(store_id, 0) if isinstance(residues, dict) else 0
        return _parse_iek_quantity(value)

    orig_rows: list[list[str]] = []
    msk_rows: list[list[str]] = []
    nsk_rows: list[list[str]] = []
    orc_rows: list[list[str]] = []
    msk_keys = list(iek_settings.get("msk_stores", []))
    nsk_key = iek_settings.get("nsk_store")
    orc_stores = list(iek_settings.get("orc_stores", []))
    prefix = str(iek_settings.get("prefix", ""))
    date_text = now.strftime(config.get("processing", {}).get("date_format", "%Y-%m-%d %H:%M"))

    for row in items:
        if not isinstance(row, dict):
            continue
        sku = str(row.get("sku") or "").strip()
        if not sku:
            continue
        sherbinka = _get_residue(row, "sherbinka")
        chehov = _get_residue(row, "chehov")
        novosibirsk = _get_residue(row, "novosibirsk")
        yasnogorsky = _get_residue(row, "yasnogorsky")
        dmd = _get_residue(row, "dmd")
        orig_rows.append([
            sku,
            str(sherbinka),
            str(chehov),
            str(novosibirsk),
            str(yasnogorsky),
            str(dmd),
        ])

        msk_value = sum(_get_residue(row, key) for key in msk_keys)
        if msk_value > 0:
            msk_rows.append([sku, str(msk_value)])

        if nsk_key:
            nsk_value = _get_residue(row, nsk_key)
            if nsk_value > 0:
                nsk_rows.append([sku, str(nsk_value)])

        for entry in orc_stores:
            if not isinstance(entry, dict):
                continue
            orc_key = entry.get("key")
            stor = entry.get("stor")
            if not orc_key or not stor:
                continue
            quantity = _get_residue(row, orc_key)
            if quantity <= 0:
                continue
            orc_rows.append([f"{prefix}{sku}", str(stor), str(quantity), date_text])

    outputs: list[Dict[str, Any]] = []
    output_names = iek_settings.get("outputs", {})
    orig_output = _render_iek_output_template(str(output_names.get("orig", "")), now, source_id)
    if orig_output:
        output_path = _resolve_output_path(file_path.parent, orig_output, _detect_output_format(orig_output))
        _write_output_file(
            output_path,
            output_path.suffix.lstrip(".").lower(),
            ["Art", "sherbinka", "chehov", "novosibirsk", "yasnogorsky", "dmd"],
            orig_rows,
        )
        outputs.append({"output": str(output_path), "rows": len(orig_rows)})

    msk_output = _render_iek_output_template(str(output_names.get("msk", "")), now, source_id)
    if msk_output:
        output_path = _resolve_output_path(file_path.parent, msk_output, _detect_output_format(msk_output))
        _write_output_file(
            output_path,
            output_path.suffix.lstrip(".").lower(),
            ["Art.", "Quant."],
            msk_rows,
        )
        outputs.append({"output": str(output_path), "rows": len(msk_rows)})

    nsk_output = _render_iek_output_template(str(output_names.get("nsk", "")), now, source_id)
    if nsk_output:
        output_path = _resolve_output_path(file_path.parent, nsk_output, _detect_output_format(nsk_output))
        _write_output_file(
            output_path,
            output_path.suffix.lstrip(".").lower(),
            ["Art.", "Quant."],
            nsk_rows,
        )
        outputs.append({"output": str(output_path), "rows": len(nsk_rows)})

    orc_output = _render_iek_output_template(str(output_names.get("orc", "")), now, source_id)
    if orc_output:
        output_path = _resolve_output_path(file_path.parent, orc_output, _detect_output_format(orc_output))
        _write_output_file(
            output_path,
            output_path.suffix.lstrip(".").lower(),
            ["Art", "Stor", "Quant", "Date"],
            orc_rows,
        )
        outputs.append({"orc_output": str(output_path), "rows": len(orc_rows)})

    results = [{"status": "success", "outputs": outputs}]
    transfer_result = _transfer_processed_outputs(
        results,
        config,
        source_id,
        source_kind,
        original_path,
        now,
        skip_original_transfer=True,
    )
    if original_path and original_path.exists():
        archive_dir = _resolve_archive_dir(config, source_kind)
        if archive_dir:
            archive_path = _archive_original_file(original_path, archive_dir, now)
            if archive_path:
                transfer_result["archive_path"] = str(archive_path)
    return {
        "status": "success",
        "processor": "iek_json",
        "results": results,
        "transfer": transfer_result,
        "source_id": source_id,
        "source_kind": source_kind,
        "file": str(file_path),
    }


def _transfer_processed_outputs(
    results: list[Dict[str, Any]],
    config: Dict[str, Any],
    source_id: str | None,
    source_kind: str | None,
    original_path: Path | None,
    now: datetime,
    skip_original_transfer: bool = False,
) -> Dict[str, Any]:
    outputs, orc_outputs = _collect_processing_outputs(results)
    if not outputs and not original_path:
        _log_processing(
            "🧩 Выгрузка остатков пропущена: нет файлов для передачи (источник=%s, тип=%s)",
            source_id or "не указан",
            source_kind or "не указан",
        )
        return {"status": "skipped", "reason": "no_outputs"}
    targets, upload_subdir = _resolve_transfer_targets(config, source_id, source_kind)
    if not targets and not orc_outputs:
        _log_processing(
            "🧩 Выгрузка остатков пропущена: не заданы ресурсы выгрузки и нет ОРК файлов",
        )
        return {"status": "skipped", "reason": "no_targets"}
    _log_processing(
        "🧩 Старт выгрузки остатков: выходных=%s, ОРК=%s, оригинал=%s, ресурсов=%s, подкаталог=%s",
        len(outputs),
        len(orc_outputs),
        original_path.name if original_path else "нет",
        len(targets),
        upload_subdir or "не задан",
    )
    smbclient_path = _resolve_smbclient_path(config)
    transfer_entries = []
    transfer_base_dir = _resolve_transfer_base_dir(config, source_kind)
    output_transfer = _transfer_files_to_targets(
        outputs,
        targets,
        upload_subdir,
        "output",
        smbclient_path,
        transfer_base_dir,
    )
    transfer_entries.extend(output_transfer["items"])
    ftp_result = _upload_orc_outputs_to_ftp(orc_outputs, config.get("ftp_ork", {}))
    _cleanup_output_files(outputs, orc_outputs, output_transfer["status_map"], ftp_result)

    original_transfer = None
    archive_path = None
    if original_path and original_path.exists() and not skip_original_transfer:
        original_transfer = _transfer_files_to_targets(
            [original_path],
            targets,
            upload_subdir,
            "original",
            smbclient_path,
            transfer_base_dir,
        )
        transfer_entries.extend(original_transfer["items"])
        if original_transfer["status_map"].get(str(original_path)):
            archive_dir = _resolve_archive_dir(config, source_kind)
            if archive_dir:
                archive_path = _archive_original_file(original_path, archive_dir, now)

    return {
        "status": "success",
        "items": transfer_entries,
        "ftp_ork": ftp_result,
        "archive_path": str(archive_path) if archive_path else None,
    }


def _collect_processing_outputs(results: list[Dict[str, Any]]) -> tuple[list[Path], list[Path]]:
    outputs: list[Path] = []
    orc_outputs: list[Path] = []
    for result in results:
        if result.get("status") != "success":
            continue
        output_groups: list[dict[str, Any]] = []
        if isinstance(result.get("outputs"), list):
            output_groups.append({"outputs": result.get("outputs")})
        if isinstance(result.get("variants"), list):
            output_groups.extend(result.get("variants") or [])
        for group in output_groups:
            if group.get("status") and group.get("status") != "success":
                continue
            for output_info in group.get("outputs", []) or []:
                output_path = output_info.get("output")
                if output_path:
                    outputs.append(Path(output_path))
                orc_output = output_info.get("orc_output")
                if orc_output:
                    orc_outputs.append(Path(orc_output))

    if outputs and orc_outputs:
        return _unique_paths(outputs), _unique_paths(orc_outputs)

    fallback_outputs: list[Path] = []
    fallback_orc_outputs: list[Path] = []

    def _walk(value: Any) -> None:
        if isinstance(value, dict):
            for key, item in value.items():
                if isinstance(item, str) and item:
                    if key in ("output", "output_path", "output_file"):
                        fallback_outputs.append(Path(item))
                        continue
                    if key in ("orc_output", "orc_output_path", "orc_output_file"):
                        fallback_orc_outputs.append(Path(item))
                        continue
                _walk(item)
        elif isinstance(value, list):
            for item in value:
                _walk(item)

    _walk(results)

    if fallback_outputs:
        outputs.extend(fallback_outputs)
    if fallback_orc_outputs:
        orc_outputs.extend(fallback_orc_outputs)

    return _unique_paths(outputs), _unique_paths(orc_outputs)


def _unique_paths(paths: Iterable[Path]) -> list[Path]:
    seen = set()
    result: list[Path] = []
    for item in paths:
        key = str(item)
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _is_unc_path(path: str) -> bool:
    path = str(path or "")
    return path.startswith("\\\\") or path.startswith("//")


def _resolve_smbclient_path(config: Dict[str, Any]) -> str | None:
    raw_path = str(config.get("smbclient_path") or "").strip()
    if raw_path:
        return raw_path
    return shutil.which("smbclient")


def _split_unc_path(unc_path: str) -> tuple[str, str, str] | None:
    unc_path = str(unc_path or "").strip()
    if not _is_unc_path(unc_path):
        return None
    windows_path = PureWindowsPath(unc_path)
    anchor = windows_path.anchor.strip("\\")
    if "\\" not in anchor:
        return None
    server, share = anchor.split("\\", 1)
    subdir = "/".join(windows_path.parts[1:])
    return server, share, subdir


def _upload_file_via_smbclient(
    file_path: Path,
    unc_path: str,
    upload_subdir: str,
    target_name: str,
    login: str | None,
    password: str | None,
    smbclient_path: str | None,
) -> Dict[str, Any]:
    if not smbclient_path:
        return {"target": target_name, "status": "error", "error": "smbclient_not_found"}
    parsed = _split_unc_path(unc_path)
    if not parsed:
        _log_processing("🧩 Некорректный UNC путь %s для выгрузки %s", unc_path, file_path.name)
        return {"target": target_name, "status": "error", "error": "invalid_unc_path"}
    server, share, base_subdir = parsed
    cleaned_subdir = _normalize_subdir(upload_subdir)
    target_subdir = _normalize_subdir("/".join(filter(None, [base_subdir, cleaned_subdir])))
    commands = []
    if target_subdir:
        subdir_parts = [part for part in target_subdir.split("/") if part]
        for part in subdir_parts:
            commands.append(f'mkdir "{part}"')
            commands.append(f'cd "{part}"')
    commands.append(f'put "{file_path}" "{file_path.name}"')
    command_text = "; ".join(commands)
    args = [smbclient_path, f"//{server}/{share}"]
    auth_file = None
    if login:
        domain, user = _split_smb_login(login)
        auth_user = user or login
        auth_file = _create_smbclient_auth_file(auth_user, password, domain)
        args.extend(["-A", auth_file])
    else:
        args.append("-N")
    args.extend(["-c", command_text])
    try:
        result = subprocess.run(args, capture_output=True, text=True, check=False)
    except Exception as exc:
        _log_processing("🧩 Ошибка запуска smbclient для %s: %s", file_path.name, exc)
        return {"target": target_name, "status": "error", "error": str(exc)}
    finally:
        if auth_file:
            try:
                Path(auth_file).unlink()
            except Exception:
                pass
    output_text = "\n".join(filter(None, [result.stdout, result.stderr])).strip()
    smbclient_error = None
    if output_text:
        error_lines = []
        for line in output_text.splitlines():
            normalized = line.strip()
            if not normalized:
                continue
            if "NT_STATUS" in normalized and "NT_STATUS_OBJECT_NAME_COLLISION" not in normalized:
                error_lines.append(normalized)
                continue
            if re.search(r"\b(error|failed|fail)\b", normalized, re.IGNORECASE):
                error_lines.append(normalized)
        if error_lines:
            smbclient_error = "; ".join(error_lines)
    if result.returncode != 0 or smbclient_error:
        _log_processing(
            "🧩 Ошибка выгрузки %s через smbclient в //%s/%s: %s",
            file_path.name,
            server,
            share,
            smbclient_error or output_text or "unknown_error",
        )
        return {
            "target": f"//{server}/{share}/{target_subdir}".rstrip("/"),
            "status": "error",
            "error": smbclient_error or output_text or "unknown_error",
        }
    _log_processing(
        "🧩 Выгружен файл %s через smbclient в //%s/%s/%s",
        file_path.name,
        server,
        share,
        target_subdir,
    )
    return {
        "target": f"//{server}/{share}/{target_subdir}".rstrip("/"),
        "status": "success",
    }


def _copy_file_to_target_dir(
    file_path: Path,
    target_dir: Path,
) -> Dict[str, Any]:
    try:
        target_dir.mkdir(parents=True, exist_ok=True)
        target_file = target_dir / file_path.name
        shutil.copy2(file_path, target_file)
        _log_processing("🧩 Выгружен файл %s в %s", file_path.name, target_dir)
        return {
            "target": str(target_dir),
            "status": "success",
        }
    except Exception as exc:
        _log_processing(
            "🧩 Ошибка переноса файла %s в %s: %s",
            file_path.name,
            target_dir,
            exc,
        )
        return {
            "target": str(target_dir),
            "status": "error",
            "error": str(exc),
        }


def _resolve_local_path_from_unc(unc_path: str) -> Path | None:
    cleaned = str(unc_path or "").replace("\\", "/")
    local_path = Path(cleaned)
    if local_path.exists():
        return local_path
    return None


def _split_smb_login(login: str) -> tuple[str | None, str | None]:
    normalized = str(login or "").strip()
    if not normalized:
        return None, None
    if "\\" in normalized:
        domain, user = normalized.split("\\", 1)
        return domain or None, user or None
    if "/" in normalized:
        domain, user = normalized.split("/", 1)
        return domain or None, user or None
    if "@" in normalized:
        user, domain = normalized.split("@", 1)
        return domain or None, user or None
    return None, normalized


def _create_smbclient_auth_file(
    login: str,
    password: str | None,
    domain: str | None,
) -> str:
    password_text = "" if password is None else str(password)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", delete=False) as handle:
        handle.write(f"username={login}\n")
        handle.write(f"password={password_text}\n")
        if domain:
            handle.write(f"domain={domain}\n")
    return handle.name


def _transfer_files_to_targets(
    files: list[Path],
    targets: list[Dict[str, Any]],
    upload_subdir: str,
    label: str,
    smbclient_path: str | None,
    base_dir: Path | None = None,
) -> Dict[str, Any]:
    transfer_entries: list[Dict[str, Any]] = []
    status_map: Dict[str, bool] = {}
    effective_subdir = _build_upload_subdir(upload_subdir)
    for file_path in files:
        if not file_path.exists():
            transfer_entries.append(
                {
                    "file": str(file_path),
                    "status": "error",
                    "error": "file_not_found",
                    "label": label,
                }
            )
            status_map[str(file_path)] = False
            continue
        target_results = []
        relative_subdir = ""
        if base_dir:
            try:
                relative_subdir = _normalize_subdir(str(file_path.parent.relative_to(base_dir)))
            except ValueError:
                relative_subdir = ""
        combined_subdir = _normalize_subdir("/".join(filter(None, [effective_subdir, relative_subdir])))
        for target in targets:
            unc_path = target.get("unc_path") or ""
            target_name = target.get("name") or target.get("id") or "resource"
            if _is_unc_path(unc_path) and os.name != "nt":
                if smbclient_path:
                    target_results.append(
                        _upload_file_via_smbclient(
                            file_path=file_path,
                            unc_path=unc_path,
                            upload_subdir=combined_subdir,
                            target_name=target_name,
                            login=target.get("login"),
                            password=target.get("password"),
                            smbclient_path=smbclient_path,
                        )
                    )
                    continue
                local_unc_path = _resolve_local_path_from_unc(unc_path)
                if local_unc_path:
                    cleaned_subdir = str(combined_subdir or "").strip().lstrip("/\\")
                    target_dir = local_unc_path / cleaned_subdir if cleaned_subdir else local_unc_path
                    target_results.append(_copy_file_to_target_dir(file_path, target_dir))
                    continue
                _log_processing(
                    "🧩 Пропуск выгрузки %s: не найден smbclient для ресурса %s",
                    file_path.name,
                    target_name,
                )
                target_results.append(
                    {
                        "target": target_name,
                        "status": "error",
                        "error": "smbclient_not_found",
                    }
                )
                continue
            target_dir = _compose_target_dir(unc_path, combined_subdir)
            if not target_dir:
                _log_processing(
                    "🧩 Пропуск выгрузки %s: не удалось определить UNC путь для ресурса %s",
                    file_path.name,
                    target_name,
                )
                target_results.append(
                    {
                        "target": target_name,
                        "status": "error",
                        "error": "empty_target",
                    }
                )
                continue
            target_results.append(_copy_file_to_target_dir(file_path, target_dir))
        success = bool(target_results) and all(item["status"] == "success" for item in target_results)
        status_map[str(file_path)] = success
        transfer_entries.append(
            {
                "file": str(file_path),
                "targets": target_results,
                "label": label,
            }
        )
    return {"items": transfer_entries, "status_map": status_map}


def _resolve_transfer_base_dir(config: Dict[str, Any], source_kind: str | None) -> Path | None:
    if source_kind == "download":
        return Path(config.get("download", {}).get("temp_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["download"]["temp_dir"]))
    if source_kind == "mail":
        return Path(config.get("mail", {}).get("temp_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["mail"]["temp_dir"]))
    return None


def _normalize_subdir(raw_subdir: str) -> str:
    cleaned = str(raw_subdir or "").strip().strip("/\\")
    if not cleaned:
        return ""
    parts = [part for part in re.split(r"[\\/]+", cleaned) if part and part != "."]
    return "/".join(parts)


def _build_upload_subdir(base_subdir: str) -> str:
    return _normalize_subdir(base_subdir)


def _resolve_source_name(
    config: Dict[str, Any],
    source_id: str | None,
    source_kind: str | None,
) -> str | None:
    if not source_id:
        return None
    sources: list[Dict[str, Any]] = []
    if source_kind == "download":
        sources = config.get("download", {}).get("sources", [])
    elif source_kind == "mail":
        sources = config.get("mail", {}).get("sources", [])
    else:
        sources = [
            *config.get("download", {}).get("sources", []),
            *config.get("mail", {}).get("sources", []),
        ]
    for item in sources:
        if str(item.get("id")) == str(source_id) or str(item.get("name")) == str(source_id):
            name = str(item.get("name") or "").strip()
            return name or None
    return None


def _resolve_source_entry(
    config: Dict[str, Any],
    source_id: str | None,
    source_kind: str | None,
) -> Dict[str, Any] | None:
    if not source_id:
        return None
    sources: list[Dict[str, Any]] = []
    if source_kind == "download":
        sources = config.get("download", {}).get("sources", [])
    elif source_kind == "mail":
        sources = config.get("mail", {}).get("sources", [])
    else:
        sources = [
            *config.get("download", {}).get("sources", []),
            *config.get("mail", {}).get("sources", []),
        ]
    for item in sources:
        if str(item.get("id")) == str(source_id) or str(item.get("name")) == str(source_id):
            return item
    return None


def _resolve_archive_dir(config: Dict[str, Any], source_kind: str | None) -> Path | None:
    if source_kind == "download":
        archive_dir = config.get("download", {}).get("archive_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["download"]["archive_dir"])
    elif source_kind == "mail":
        archive_dir = config.get("mail", {}).get("archive_dir", DEFAULT_SUPPLIER_STOCK_CONFIG["mail"]["archive_dir"])
    else:
        return None
    return Path(archive_dir)

def _parse_bool_config(value: Any, default: bool) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in {"1", "true", "yes", "y", "on", "да"}:
            return True
        if normalized in {"0", "false", "no", "n", "off", "нет"}:
            return False
    return bool(value)

def _is_passive_mode_denied_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return "passive mode denied" in message or "pasv" in message or message.startswith("451 ")

def _upload_file_to_ftp(ftp: ftplib.FTP, orc_path: Path, remote_name: str | None = None) -> None:
    target_name = remote_name or orc_path.name
    with orc_path.open("rb") as handle:
        ftp.storbinary(f"STOR {target_name}", handle)

def _upload_orc_outputs_to_ftp(orc_outputs: list[Path], ftp_config: Dict[str, Any]) -> Dict[str, Any]:
    if not orc_outputs:
        _log_processing("🧩 FTP ОРК: выгрузка пропущена (нет файлов)")
        return {"status": "skipped", "reason": "no_orc_files", "items": []}
    host_raw = str(ftp_config.get("host") or "").strip()
    if not host_raw:
        _log_processing("🧩 FTP ОРК: выгрузка пропущена (хост не задан)")
        return {"status": "skipped", "reason": "no_host", "items": []}
    host, port, base_dir = _parse_ftp_host(host_raw)
    login = str(ftp_config.get("login") or "").strip() or None
    password = str(ftp_config.get("password") or "") if ftp_config.get("password") is not None else ""
    passive = _parse_bool_config(ftp_config.get("passive"), True)
    results: list[Dict[str, Any]] = []
    try:
        with ftplib.FTP() as ftp:
            ftp.connect(host, port, timeout=30)
            if login:
                ftp.login(login, password)
            else:
                ftp.login()
            ftp.set_pasv(passive)
            if base_dir:
                try:
                    ftp.cwd(base_dir)
                except Exception as exc:
                    _log_processing("🧩 Ошибка перехода в каталог FTP ОРК %s: %s", base_dir, exc)
                    return {"status": "error", "error": str(exc), "items": results}
            for orc_path in orc_outputs:
                if not orc_path.exists():
                    results.append({"file": str(orc_path), "status": "error", "error": "file_not_found"})
                    continue
                remote_name = "iek.csv" if orc_path.name == "iek_4ork.csv" else orc_path.name
                try:
                    _upload_file_to_ftp(ftp, orc_path, remote_name)
                    _log_processing("🧩 Выгружен ОРК файл %s на FTP %s", remote_name, host)
                    results.append({"file": str(orc_path), "status": "success", "remote_name": remote_name})
                except Exception as exc:
                    if passive and _is_passive_mode_denied_error(exc):
                        _log_processing(
                            "🧩 Пассивный режим FTP запрещен, повторяем в активном режиме для %s",
                            remote_name,
                        )
                        passive = False
                        ftp.set_pasv(False)
                        try:
                            _upload_file_to_ftp(ftp, orc_path, remote_name)
                            _log_processing(
                                "🧩 Выгружен ОРК файл %s на FTP %s (активный режим)",
                                remote_name,
                                host,
                            )
                            results.append({"file": str(orc_path), "status": "success", "remote_name": remote_name})
                        except Exception as retry_exc:
                            _log_processing("🧩 Ошибка выгрузки ОРК %s по FTP: %s", remote_name, retry_exc)
                            results.append({"file": str(orc_path), "status": "error", "error": str(retry_exc)})
                    else:
                        _log_processing("🧩 Ошибка выгрузки ОРК %s по FTP: %s", remote_name, exc)
                        results.append({"file": str(orc_path), "status": "error", "error": str(exc)})
    except Exception as exc:
        _log_processing("🧩 Ошибка подключения к FTP ОРК: %s", exc)
        return {"status": "error", "error": str(exc), "items": results}
    return {"status": "success", "items": results}


def _parse_ftp_host(host_raw: str) -> tuple[str, int, str | None]:
    host_raw = host_raw.strip()
    base_dir = None
    if "://" in host_raw:
        parsed = urlparse(host_raw)
        if parsed.hostname:
            base_dir = parsed.path.strip("/") or None
            return parsed.hostname, parsed.port or 21, base_dir
    if "/" in host_raw:
        host_raw, base_dir = host_raw.split("/", 1)
        base_dir = base_dir.strip("/") or None
    if ":" in host_raw:
        host, port_text = host_raw.rsplit(":", 1)
        if port_text.isdigit():
            return host, int(port_text), base_dir
    return host_raw, 21, base_dir


def _cleanup_output_files(
    outputs: list[Path],
    orc_outputs: list[Path],
    status_map: Dict[str, bool],
    ftp_result: Dict[str, Any],
) -> None:
    ftp_status_map = {}
    if ftp_result.get("status") == "success":
        for item in ftp_result.get("items", []):
            ftp_status_map[str(item.get("file"))] = item.get("status") == "success"
    for output_path in _unique_paths(outputs + orc_outputs):
        output_key = str(output_path)
        if output_path in orc_outputs:
            if not ftp_status_map.get(output_key):
                continue
        elif not status_map.get(output_key):
            continue
        try:
            output_path.unlink()
        except Exception as exc:
            _log_processing(
                "🧩 Не удалось удалить исходный файл %s после выгрузки: %s",
                output_path.name,
                exc,
            )


def _resolve_transfer_targets(
    config: Dict[str, Any],
    source_id: str | None,
    source_kind: str | None,
) -> tuple[list[Dict[str, Any]], str]:
    sources: list[Dict[str, Any]] = []
    source = None
    if source_kind == "download":
        sources = config.get("download", {}).get("sources", [])
    elif source_kind == "mail":
        sources = config.get("mail", {}).get("sources", [])
    else:
        sources = [
            *config.get("download", {}).get("sources", []),
            *config.get("mail", {}).get("sources", []),
        ]
    if source_id:
        for item in sources:
            if str(item.get("id")) == str(source_id) or str(item.get("name")) == str(source_id):
                source = item
                break
    upload_subdir = str(source.get("upload_subdir") or "").strip() if source else ""
    individual = source.get("individual_directory", {}) if source else {}
    if isinstance(individual, dict) and individual.get("enabled"):
        individual_path = str(individual.get("unc_path") or "").strip()
        if individual_path:
            return (
                [
                    {
                        "id": "individual",
                        "name": "Индивидуальный каталог",
                        "unc_path": individual_path,
                        "login": individual.get("login"),
                        "password": individual.get("password"),
                    }
                ],
                upload_subdir,
            )
    resources = config.get("resources", [])
    targets = []
    if isinstance(resources, list):
        for resource in resources:
            if not isinstance(resource, dict):
                continue
            if not resource.get("enabled", True):
                continue
            unc_path = str(resource.get("unc_path") or "").strip()
            if not unc_path:
                continue
            targets.append(
                {
                    "id": resource.get("id") or resource.get("name"),
                    "name": resource.get("name") or resource.get("id"),
                    "unc_path": unc_path,
                    "login": resource.get("login"),
                    "password": resource.get("password"),
                }
            )
    return targets, upload_subdir


def _compose_target_dir(base_path: str, subdir: str) -> Path | None:
    base_path = str(base_path or "").strip()
    if not base_path:
        return None
    cleaned_subdir = str(subdir or "").strip().lstrip("/\\")
    if _is_unc_path(base_path):
        if os.name != "nt":
            _log_processing(
                "🧩 UNC путь %s задан, но текущая ОС не Windows. Выгрузка пропущена.",
                base_path,
            )
            return None
        unc_path = PureWindowsPath(base_path)
        if cleaned_subdir:
            return Path(unc_path / cleaned_subdir)
        return Path(unc_path)
    if cleaned_subdir:
        return Path(base_path) / cleaned_subdir
    return Path(base_path)


def _processing_rule_matches_kind(
    rule: Dict[str, Any],
    source_kind: str,
    config: Dict[str, Any],
) -> bool:
    rule_kind = rule.get("source_kind")
    if rule_kind:
        return rule_kind == source_kind
    resolved_kind = _resolve_processing_rule_source_kind(rule, config)
    return resolved_kind == source_kind


def _resolve_processing_rule_source_kind(
    rule: Dict[str, Any],
    config: Dict[str, Any],
) -> str | None:
    source_id = rule.get("source_id")
    if not source_id:
        return None
    download_sources = config.get("download", {}).get("sources", [])
    mail_sources = config.get("mail", {}).get("sources", [])
    download_source = next(
        (item for item in download_sources if str(item.get("id")) == str(source_id)),
        None,
    )
    mail_source = next(
        (item for item in mail_sources if str(item.get("id")) == str(source_id)),
        None,
    )
    if download_source and not mail_source:
        return "download"
    if mail_source and not download_source:
        return "mail"
    if download_source and mail_source:
        source_file = str(rule.get("source_file") or "")
        if source_file and source_file == str(download_source.get("output_name") or ""):
            return "download"
        if source_file and source_file == str(mail_source.get("output_template") or ""):
            return "mail"
        rule_name = str(rule.get("name") or "")
        if rule_name and rule_name == str(download_source.get("name") or ""):
            return "download"
        if rule_name and rule_name == str(mail_source.get("name") or ""):
            return "mail"
    return None


def _processing_rule_matches(
    rule: Dict[str, Any],
    file_path: Path,
    input_index: int | None,
) -> bool:
    source_file = str(rule.get("source_file") or "").strip()
    if not source_file:
        return False
    target_name = file_path.name
    if source_file == target_name:
        return True
    rendered_source = _render_output_name_template(source_file, file_path, input_index)
    if rendered_source and rendered_source == target_name:
        return True
    if _processing_rule_matches_unpacked_archive(rendered_source or source_file, file_path):
        return True
    template_pattern = _processing_rule_template_to_glob(source_file)
    if template_pattern != source_file and fnmatch(target_name, template_pattern):
        return True
    if any(char in source_file for char in ("*", "?", "[")):
        return fnmatch(target_name, source_file)
    return False


def _processing_rule_template_to_glob(source_file: str) -> str:
    def _replace(match: re.Match[str]) -> str:
        key = match.group("key")
        if key in ("index", "name", "filename"):
            return "*"
        return match.group(0)

    return re.sub(r"\{(?P<key>[A-Za-z_][A-Za-z0-9_]*)(?::[^}]*)?\}", _replace, source_file)


def _processing_rule_matches_unpacked_archive(source_name: str, file_path: Path) -> bool:
    if not source_name:
        return False
    name = Path(str(source_name)).name
    if not _is_archive_name(name):
        return False
    base_name = _strip_archive_suffix(Path(name))
    if not base_name:
        return False
    return file_path.stem == base_name or file_path.name == base_name


def _run_processing_rule(
    file_path: Path,
    rule: Dict[str, Any],
    processing: Dict[str, Any],
    now: datetime,
    input_index: int | None,
    source_name: str | None = None,
) -> Dict[str, Any]:
    if not _normalize_requires_processing(rule.get("requires_processing", True)):
        _logger.info("🧩 Обработка не требуется для %s (%s)", file_path, rule.get("name") or rule.get("id"))
        return {"status": "skipped", "reason": "processing_disabled", "rule_id": rule.get("id")}

    table = _read_supplier_table(file_path)
    if not table:
        return {"status": "error", "error": "empty_table", "rule_id": rule.get("id")}

    data_row = int(rule.get("data_row") or 1)
    variants = rule.get("variants", [])
    results = []
    for variant in variants:
        variant_result = _process_variant(
            table,
            file_path,
            rule,
            variant,
            processing,
            now,
            data_row,
            input_index,
            source_name,
        )
        results.append(variant_result)

    return {"status": "success", "rule_id": rule.get("id"), "variants": results}


def _read_supplier_table(file_path: Path) -> list[list[str]]:
    suffix = file_path.suffix.lower().lstrip(".")
    if suffix == "csv":
        return _read_csv_table(file_path)
    if suffix in ("xlsx",):
        return _read_xlsx_table(file_path)
    if suffix in ("xls",):
        return _read_xls_table(file_path)
    raise ValueError(f"Неизвестный формат файла: {file_path.suffix}")


def _read_csv_table(file_path: Path) -> list[list[str]]:
    with file_path.open("r", encoding="utf-8", errors="ignore", newline="") as file:
        sample = file.read(2048)
        file.seek(0)
        delimiter = _guess_csv_delimiter(sample)
        reader = csv.reader(file, delimiter=delimiter)
        return [[str(value) for value in row] for row in reader]


def _guess_csv_delimiter(sample: str) -> str:
    if not sample:
        return ","
    comma = sample.count(",")
    semicolon = sample.count(";")
    tab = sample.count("\t")
    if semicolon > comma and semicolon >= tab:
        return ";"
    if tab > comma and tab > semicolon:
        return "\t"
    return ","


def _read_xlsx_table(file_path: Path) -> list[list[str]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl не установлен для обработки xlsx") from exc
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active
    rows: list[list[str]] = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([_normalize_cell(value) for value in row])
    return rows


def _read_xls_table(file_path: Path) -> list[list[str]]:
    try:
        import xlrd
    except ImportError as exc:
        if _looks_like_xlsx(file_path):
            _logger.warning(
                "⚠️ Файл %s имеет расширение .xls, но похож на xlsx. Используем openpyxl.",
                file_path.name,
            )
            return _read_xlsx_table(file_path)
        raise ImportError(
            "xlrd не установлен для обработки xls. Установите пакет xlrd или сохраните файл в формате xlsx."
        ) from exc
    try:
        workbook = xlrd.open_workbook(file_path)
    except xlrd.biffh.XLRDError:
        if _looks_like_xlsx(file_path):
            _logger.warning(
                "⚠️ Файл %s имеет расширение .xls, но похож на xlsx. Используем openpyxl.",
                file_path.name,
            )
            return _read_xlsx_table(file_path)
        raise
    sheet = workbook.sheet_by_index(0)
    rows: list[list[str]] = []
    for row_index in range(sheet.nrows):
        rows.append([_normalize_cell(sheet.cell_value(row_index, col)) for col in range(sheet.ncols)])
    return rows


def _looks_like_xlsx(file_path: Path) -> bool:
    try:
        with file_path.open("rb") as file:
            signature = file.read(4)
    except OSError:
        return False
    return signature in (b"PK\x03\x04", b"PK\x05\x06", b"PK\x07\x08")


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isfinite(value) and value.is_integer():
            return str(int(value))
    return str(value)


def _normalize_spacing_tokens(value: str) -> str:
    if not value:
        return ""
    return value.replace("\\s", " ").replace("\\t", "\t")


def _process_variant(
    table: list[list[str]],
    file_path: Path,
    rule: Dict[str, Any],
    variant: Dict[str, Any],
    processing: Dict[str, Any],
    now: datetime,
    data_row: int,
    input_index: int | None,
    source_name: str | None,
) -> Dict[str, Any]:
    article_col = int(variant.get("article_col") or 0)
    if article_col <= 0:
        return {"status": "error", "error": "invalid_article_col"}
    data_columns = variant.get("data_columns", [])
    output_names = variant.get("output_names", [])
    output_format = (variant.get("output_format") or "csv").lower()
    article_filter = variant.get("article_filter") or ""
    extra_filter = variant.get("extra_filter") or ""
    extra_filter_col = int(variant.get("extra_filter_col") or 0)
    use_article_filter = variant.get("use_article_filter")
    if use_article_filter is None:
        use_article_filter = bool(article_filter)
    use_article_filter_columns = list(variant.get("use_article_filter_columns", []))
    if len(use_article_filter_columns) < len(data_columns):
        use_article_filter_columns.extend([True] * (len(data_columns) - len(use_article_filter_columns)))
    use_article_filter_columns = use_article_filter_columns[:len(data_columns)]
    article_prefix = _normalize_spacing_tokens(variant.get("article_prefix") or "")
    article_postfix = variant.get("article_postfix") or ""
    article_transform = variant.get("article_transform") or {}
    transform_pattern = ""
    transform_replacement = ""
    if isinstance(article_transform, dict):
        transform_pattern = article_transform.get("pattern") or ""
        transform_replacement = article_transform.get("replacement") or ""
    orc_config = variant.get("orc", {}) if isinstance(variant.get("orc"), dict) else {}
    orc_enabled = bool(orc_config.get("enabled"))
    orc_column = int(orc_config.get("column") or 0)
    orc_output_index = int(orc_config.get("output_index") or 0)
    if 1 <= orc_output_index <= len(data_columns):
        orc_output_index_effective = orc_output_index
    elif data_columns:
        orc_output_index_effective = 1
    else:
        orc_output_index_effective = 0
    orc_input_index = int(orc_config.get("input_index") or 0)
    input_index_value = input_index or 1
    orc_input_match = orc_input_index <= 0 or orc_input_index == input_index_value

    if len(data_columns) != len(output_names):
        return {"status": "error", "error": "columns_names_mismatch"}

    compiled_filter = None
    if use_article_filter and article_filter:
        try:
            compiled_filter = re.compile(article_filter)
        except re.error as exc:
            return {"status": "error", "error": f"invalid_filter: {exc}"}
    compiled_extra_filter = None
    if extra_filter and extra_filter_col > 0:
        try:
            compiled_extra_filter = re.compile(extra_filter)
        except re.error as exc:
            return {"status": "error", "error": f"invalid_extra_filter: {exc}"}
    elif extra_filter:
        return {"status": "error", "error": "invalid_extra_filter_col"}
    compiled_transform = None
    if transform_pattern:
        try:
            compiled_transform = re.compile(transform_pattern)
        except re.error as exc:
            return {"status": "error", "error": f"invalid_article_transform: {exc}"}

    rows = table[data_row - 1:] if data_row > 1 else table
    outputs: list[Dict[str, Any]] = []
    orc_output_written = False
    for idx, (column_index, output_name) in enumerate(zip(data_columns, output_names)):
        input_template = str(rule.get("source_file") or "").strip() or None
        render_context = {
            "source_id": rule.get("source_id"),
            "source_name": source_name,
        }
        if source_name:
            render_context["name"] = source_name
        rendered_output_name = _render_output_name_template(
            output_name,
            file_path,
            input_index,
            input_template,
            render_context,
        )
        rendered_output_name = _apply_input_index_to_output_name(
            rendered_output_name,
            str(output_name),
            input_index,
        )
        items: list[list[str]] = []
        orc_items: list[list[str]] = []
        orc_active = (
            orc_enabled
            and orc_input_match
            and not orc_output_written
            and orc_output_index_effective
            and idx + 1 == orc_output_index_effective
        )
        use_filter_for_column = use_article_filter and (
            use_article_filter_columns[idx] if idx < len(use_article_filter_columns) else True
        )
        for row in rows:
            article_raw = _get_cell(row, article_col, preserve_whitespace=True)
            article_source = article_raw
            article_transformed = (
                compiled_transform.sub(transform_replacement, article_source)
                if compiled_transform
                else article_source
            )
            article_trimmed = article_transformed.strip()
            if not article_trimmed:
                continue
            if compiled_filter and use_filter_for_column and not compiled_filter.search(article_trimmed):
                continue
            if compiled_extra_filter:
                extra_value = _get_cell(row, extra_filter_col, preserve_whitespace=True).strip()
                if not compiled_extra_filter.search(extra_value):
                    continue
            quant_raw = _get_cell(row, column_index)
            quant_value = _parse_quantity(quant_raw)
            if quant_value is None:
                continue
            article_value = _apply_article_postfix(f"{article_prefix}{article_transformed}", article_postfix)
            items.append([article_value, quant_value])
            if orc_active:
                orc_prefix = _normalize_spacing_tokens(orc_config.get("prefix", ""))
                stor = orc_config.get("stor", "")
                orc_column_index = orc_column if orc_column > 0 else column_index
                orc_quant_raw = _get_cell(row, orc_column_index)
                orc_quant_value = _parse_quantity(orc_quant_raw)
                if orc_quant_value is None:
                    continue
                date_text = now.strftime(processing.get("date_format", "%Y-%m-%d %H:%M"))
                orc_items.append([f"{orc_prefix}{article_transformed}", stor, orc_quant_value, date_text])

        output_path = _resolve_output_path(file_path.parent, rendered_output_name, output_format)
        output_path = _write_output_file(output_path, output_format, ["Art.", "Quant."], items)

        orc_output = None
        if orc_active:
            orc_output_format = (orc_config.get("output_format") or output_format).lower()
            orc_output_name = str(orc_config.get("output_name") or "").strip()
            if orc_output_name:
                rendered_orc_name = _render_output_name_template(
                    orc_output_name,
                    file_path,
                    input_index,
                    input_template,
                    render_context,
                )
            else:
                rendered_orc_name = _append_suffix_to_name(rendered_output_name, "_orc")
            orc_output_path = _resolve_output_path(
                file_path.parent,
                rendered_orc_name,
                orc_output_format,
            )
            orc_output_path = _write_output_file(
                orc_output_path,
                orc_output_format,
                ["Art", "Stor", "Quant", "Date"],
                orc_items,
            )
            orc_output = str(orc_output_path)
            orc_output_written = True

        outputs.append(
            {
                "output": str(output_path),
                "orc_output": orc_output,
                "rows": len(items),
            }
        )

    return {"status": "success", "outputs": outputs}


def _get_cell(row: list[str], index: int, *, preserve_whitespace: bool = False) -> str:
    if index <= 0:
        return ""
    if index - 1 >= len(row):
        return ""
    value = row[index - 1]
    if value is None:
        return ""
    text = str(value)
    if preserve_whitespace:
        return text
    return text.strip()


def _parse_quantity(value: str) -> str | None:
    if value is None:
        return None
    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    if number <= 0:
        return None
    if number.is_integer():
        return str(int(number))
    return str(number)


def _parse_iek_quantity(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return max(0, value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return 0
        return max(0, int(value))
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return 0
    try:
        number = float(text)
    except ValueError:
        return 0
    if not math.isfinite(number) or number <= 0:
        return 0
    return int(number)


def _apply_article_postfix(article: str, postfix: str) -> str:
    if not postfix:
        return article
    if postfix[0].isspace() or article.endswith((" ", "\t")):
        return f"{article}{postfix}"
    return f"{article} {postfix}"


def _resolve_output_path(folder: Path, output_name: str, output_format: str) -> Path:
    name = output_name.strip()
    suffix = f".{output_format}"
    if not name:
        name = f"output{suffix}"
    if Path(name).suffix.lower() != suffix:
        name = f"{name}{suffix}"
    return folder / name


def _append_suffix_to_name(filename: str, suffix: str) -> str:
    path = Path(filename)
    return f"{path.stem}{suffix}{path.suffix}"


def _match_template_filename(template: str, filename: str) -> Dict[str, str]:
    if not template:
        return {}
    placeholder_pattern = r"\{(?P<key>name|index|filename)(?::[^}]*)?\}"
    if not re.search(placeholder_pattern, template):
        return {}

    def _group_for_placeholder(match: re.Match[str]) -> str:
        key = match.group("key")
        if key == "index":
            return r"(?P<index>\d+)"
        return rf"(?P<{key}>.+?)"

    pattern_parts = []
    last_index = 0
    for match in re.finditer(placeholder_pattern, template):
        pattern_parts.append(re.escape(template[last_index:match.start()]))
        pattern_parts.append(_group_for_placeholder(match))
        last_index = match.end()
    pattern_parts.append(re.escape(template[last_index:]))
    pattern = "".join(pattern_parts)
    match = re.match(rf"^{pattern}$", filename)
    return match.groupdict() if match else {}


def _resolve_output_name_values(
    template: str,
    file_path: Path,
    input_index: int | None,
    input_template: str | None = None,
    extra_values: Dict[str, Any] | None = None,
) -> Dict[str, Any]:
    values: Dict[str, Any] = {
        "index": input_index or 1,
        "name": file_path.stem,
        "filename": file_path.name,
    }
    locked_keys: set[str] = set()
    extracted_from_input = _match_template_filename(str(input_template or ""), file_path.name)
    if "name" in extracted_from_input:
        values["name"] = extracted_from_input["name"]
        locked_keys.add("name")
    if "filename" in extracted_from_input:
        values["filename"] = extracted_from_input["filename"]
        locked_keys.add("filename")
    if "index" in extracted_from_input:
        try:
            values["index"] = int(extracted_from_input["index"])
        except ValueError:
            values["index"] = extracted_from_input["index"]
        locked_keys.add("index")
    extracted = _match_template_filename(template, file_path.name)
    if "name" in extracted:
        values["name"] = extracted["name"]
        locked_keys.add("name")
    if "filename" in extracted:
        values["filename"] = extracted["filename"]
        locked_keys.add("filename")
    if "index" in extracted:
        try:
            values["index"] = int(extracted["index"])
        except ValueError:
            values["index"] = extracted["index"]
        locked_keys.add("index")
    if isinstance(extra_values, dict):
        for key, value in extra_values.items():
            if value is None:
                continue
            if key in locked_keys:
                continue
            values[key] = value
    return values


def _render_output_name_template(
    template: str,
    file_path: Path,
    input_index: int | None,
    input_template: str | None = None,
    extra_values: Dict[str, Any] | None = None,
) -> str:
    if not template:
        return template
    values = _resolve_output_name_values(template, file_path, input_index, input_template, extra_values)
    try:
        return template.format(**values)
    except Exception:
        return template


def _apply_input_index_to_output_name(
    rendered_name: str,
    template: str,
    input_index: int | None,
) -> str:
    if not rendered_name or not input_index or input_index <= 1:
        return rendered_name
    if re.search(r"\{(?:index|name|filename)(?::[^}]*)?\}", template):
        return rendered_name
    match = re.search(r"\d+$", rendered_name)
    if match:
        return f"{rendered_name[:match.start()]}{input_index}"
    return f"{rendered_name}{input_index}"


def _write_output_file(path: Path, fmt: str, headers: list[str], rows: list[list[str]]) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    if fmt == "csv":
        with path.open("w", encoding="utf-8", newline="") as file:
            writer = csv.writer(file, delimiter=";")
            writer.writerow(headers)
            writer.writerows(rows)
        return path
    if fmt == "xlsx":
        if _write_xlsx(path, headers, rows):
            return path
        return _write_fallback_csv(path, headers, rows, fmt)
    if fmt == "xls":
        if _write_xls(path, headers, rows):
            return path
        xlsx_path = path.with_suffix(".xlsx")
        if _write_xlsx(xlsx_path, headers, rows):
            _logger.warning("⚠️ Запись xls недоступна, сохранено как %s", xlsx_path.name)
            return xlsx_path
        return _write_fallback_csv(path, headers, rows, fmt)
    raise ValueError(f"Неизвестный формат выходного файла: {fmt}")


def _write_xlsx(path: Path, headers: list[str], rows: list[list[str]]) -> bool:
    try:
        import openpyxl
    except ImportError as exc:
        _logger.warning("openpyxl не установлен для записи xlsx (%s)", exc)
        return False
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    return True


def _write_xls(path: Path, headers: list[str], rows: list[list[str]]) -> bool:
    try:
        import xlwt
    except ImportError as exc:
        _logger.warning("xlwt не установлен для записи xls (%s)", exc)
        return False
    workbook = xlwt.Workbook()
    sheet = workbook.add_sheet("Sheet1")
    for col, header in enumerate(headers):
        sheet.write(0, col, header)
    for row_index, row in enumerate(rows, start=1):
        for col_index, value in enumerate(row):
            sheet.write(row_index, col_index, value)
    workbook.save(str(path))
    return True


def _write_fallback_csv(path: Path, headers: list[str], rows: list[list[str]], fmt: str) -> Path:
    fallback_path = path.with_suffix(".csv")
    _logger.warning("⚠️ Запись %s недоступна, сохранено как %s", fmt, fallback_path.name)
    return _write_output_file(fallback_path, "csv", headers, rows)


def _strip_archive_suffix(path: Path) -> str:
    name = path.name
    for suffix in (".tar.gz", ".tar.bz2", ".tar.xz", ".tgz", ".zip", ".tar", ".gz", ".bz2", ".xz"):
        if name.endswith(suffix):
            return name[: -len(suffix)]
    return path.stem


def _is_archive_name(name: str) -> bool:
    return name.endswith((".tar.gz", ".tar.bz2", ".tar.xz", ".tgz", ".zip", ".tar", ".gz", ".bz2", ".xz"))


def unpack_archive_file(archive_path: Path) -> Path | None:
    if not archive_path.exists():
        return None
    extracted_path: Path | None = None
    try:
        if zipfile.is_zipfile(archive_path):
            with zipfile.ZipFile(archive_path) as zip_file:
                members = [name for name in zip_file.namelist() if name and not name.endswith("/")]
                if not members:
                    return None
                member = members[0]
                extracted_path = Path(zip_file.extract(member, path=archive_path.parent))
        elif tarfile.is_tarfile(archive_path):
            with tarfile.open(archive_path) as tar_file:
                members = [member for member in tar_file.getmembers() if member.isfile()]
                if not members:
                    return None
                member = members[0]
                tar_file.extract(member, path=archive_path.parent)
                extracted_path = archive_path.parent / member.name
        else:
            return None

        if not extracted_path or not extracted_path.exists():
            return None

        base_name = _strip_archive_suffix(archive_path)
        new_name = f"{base_name}{extracted_path.suffix}" if extracted_path.suffix else base_name
        target_path = archive_path.with_name(new_name)
        if target_path.exists():
            target_path.unlink()
        extracted_path.replace(target_path)
        return target_path
    except Exception:
        return None


def summarize_supplier_stock_sources(sources: Iterable[Dict[str, Any]]) -> List[Dict[str, Any]]:
    summary = []
    for source in sources:
        summary.append(
            {
                "id": source.get("id"),
                "name": source.get("name"),
                "url": source.get("url"),
                "output_name": source.get("output_name"),
                "method": source.get("method", "http"),
                "enabled": source.get("enabled", True),
            }
        )
    return summary
